import os
import re
import smtplib
import subprocess
import sys
import tempfile
import uuid
import shutil
import html
from datetime import datetime, timedelta
from io import BytesIO
from email.message import EmailMessage
from email.utils import parseaddr

import pandas as pd
from flask import Flask, flash, jsonify, redirect, render_template, request, url_for
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

try:
    import win32com.client as win32
except ImportError:
    win32 = None

try:
    import pythoncom
except ImportError:
    pythoncom = None

app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET_KEY", "email-sender-dev-secret")
MISC_TO_RECIPIENTS = ["mmemon@orthosynetics.com"] # later change to mmemon@orthosynetics.com
MISC_CC_RECIPIENTS = ["rafiks77@gmail.com"] # later change to rafiks77@gmail.com

WORKFLOWS = [
    ("bv-ortho-no-info", "BV – Ortho No Info Emailer"),
    ("bv-dental-no-info", "BV – Dental No Info Emailer"),
    ("ev-ortho-rstc", "EV – Ortho RSTC Emailer"),
    ("ev-dental-rstc", "EV – Dental RSTC Emailer"),
    ("bv-ortho-rstc", "BV – Ortho RSTC Emailer"),
    ("bv-dental-rstc", "BV – Dental RSTC Emailer"),
]
WORKFLOW_MAP = dict(WORKFLOWS)
DEFAULT_WORKFLOW = WORKFLOWS[0][0]
NO_INFO_WORKFLOWS = {"bv-ortho-no-info", "bv-dental-no-info"}
EV_RSTC_WORKFLOWS = {"ev-ortho-rstc", "ev-dental-rstc"}
BV_PREV_DAY_WORKFLOWS = {"bv-ortho-rstc", "bv-dental-rstc"}
RSTC_HIGHLIGHT_WORKFLOWS = {
    "ev-ortho-rstc",
    "ev-dental-rstc",
    "bv-ortho-rstc",
    "bv-dental-rstc",
}
EV_PDF_UPLOAD_OFFICE_NAMES = {"FREDPEDO", "MUSGROVE", "REISTERS"}
EV_EXCLUDED_OUTPUT_COLUMNS = {
    "officename",
    "patientsname",
    "patientidchart",
    "status",
}
STATE_BY_WORKFLOW = {
    workflow_id: {
        "records": [],
        "summary": {},
        "uploaded": False,
    }
    for workflow_id, _ in WORKFLOWS
}


def get_workflow_or_default(workflow_id):
    if workflow_id in WORKFLOW_MAP:
        return workflow_id
    return DEFAULT_WORKFLOW


def reset_state(workflow_id):
    state = STATE_BY_WORKFLOW[workflow_id]
    state["records"] = []
    state["summary"] = {}
    state["uploaded"] = False


def safe_file_name(name):
    return "".join(c for c in name if c.isalnum() or c in (" ", "-", "_")).strip()


def normalize_column_name(name):
    return "".join(ch for ch in str(name).strip().lower() if ch.isalnum())


def find_column(df, expected_name):
    expected_normalized = normalize_column_name(expected_name)
    for column in df.columns:
        if normalize_column_name(column) == expected_normalized:
            return column
    raise ValueError(
        f"Required column '{expected_name}' not found. Available columns: {', '.join(map(str, df.columns))}"
    )


def find_optional_column(df, expected_name):
    try:
        return find_column(df, expected_name)
    except ValueError:
        return None


def clean_recipients(raw_value):
    if pd.isna(raw_value):
        return []
    recipients = []
    for item in str(raw_value).replace(";", ",").split(","):
        candidate = item.strip()
        if not candidate or candidate.lower() == "nan":
            continue

        _, parsed_email = parseaddr(candidate)
        email = parsed_email.strip() if parsed_email else candidate
        if "@" in email:
            recipients.append(email)
    return recipients


def workflow_uses_no_info_body_mode(workflow_id):
    return workflow_id in NO_INFO_WORKFLOWS


def should_attach_extra_pdf(workflow_id, office_name):
    return (
        workflow_id in EV_RSTC_WORKFLOWS
        and str(office_name).strip().upper() in EV_PDF_UPLOAD_OFFICE_NAMES
    )


def has_no_info_text(value):
    if pd.isna(value):
        return False
    text = str(value).strip().lower()
    if not text:
        return False
    return re.search(r"n\W*o\W*i\W*n\W*f\W*o", text) is not None


def filter_no_info_rows(df, column_names):
    mask = pd.Series(False, index=df.index)
    for column_name in column_names:
        mask = mask | df[column_name].apply(has_no_info_text)
    return df[mask]


def format_date_columns_in_df(df):
    formatted_df = df.copy()
    target_columns = {
        "appointment",
        "appoinment",
        "appointmentdate",
        "appoinmentdate",
        "apptdate",
        "dob",
        "dos",
        "dosdate",
        "dosdob",
        "DOS/DOB",
    }
    for column_name in formatted_df.columns:
        if normalize_column_name(column_name) not in target_columns:
            continue

        original_values = formatted_df[column_name]
        converted_values = []
        for original in original_values:
            if pd.isna(original):
                converted_values.append("")
                continue

            # Parse value-by-value so mixed formats (e.g. MM/DD/YYYY + ISO datetime)
            # are consistently normalized.
            parsed = pd.to_datetime(original, errors="coerce")
            if pd.notna(parsed):
                converted_values.append(parsed.strftime("%m/%d/%Y"))
            else:
                converted_values.append(str(original))

        formatted_df[column_name] = converted_values
    return formatted_df


def format_date_for_display(value):
    if pd.isna(value):
        return ""
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.notna(parsed):
        return parsed.strftime("%m/%d/%Y")
    return str(value).strip()


def split_patient_name(raw_name):
    if pd.isna(raw_name):
        return "", ""
    text = str(raw_name).strip()
    if not text:
        return "", ""
    if "," in text:
        last_name, first_name = text.split(",", 1)
        return last_name.strip(), first_name.strip()
    return text, ""


def transform_fredpedo_slice_df(df):
    patient_name_col = find_optional_column(df, "Patients Name")
    dob_col = find_optional_column(df, "DOB")
    insurance_col = find_optional_column(df, "Insurance")
    carrier_phone_col = find_optional_column(df, "Carrier Phone")
    subscriber_name_col = find_optional_column(df, "Subscriber Name")
    policy_id_col = find_optional_column(df, "Policy ID")
    subscriber_dob_col = find_optional_column(df, "Subscriber DOB")
    appointment_col = (
        find_optional_column(df, "Appointment")
        or find_optional_column(df, "Appointment Date")
        or find_optional_column(df, "Appoinment Date")
    )
    comments_col = find_optional_column(df, "Comments")

    transformed_rows = []
    for _, row in df.iterrows():
        pats_last_name, pats_first_name = split_patient_name(
            row[patient_name_col] if patient_name_col else ""
        )

        transformed_rows.append(
            {
                "PatsLastName": pats_last_name,
                "PatsFirstName": pats_first_name,
                "PatsBirthDate": row[dob_col] if dob_col else "",
                "CarrierName": row[insurance_col] if insurance_col else "",
                "CarrierPhone": row[carrier_phone_col] if carrier_phone_col else "",
                "EmpName": row[subscriber_name_col] if subscriber_name_col else "",
                "PolEmployeeSSN": row[policy_id_col] if policy_id_col else "",
                "EmployeeBirthDate": row[subscriber_dob_col] if subscriber_dob_col else "",
                "GroupName": "",
                "FutureAppt": row[appointment_col] if appointment_col else "",
                "comments": row[comments_col] if comments_col else "",
            }
        )

    return pd.DataFrame(transformed_rows)


def drop_columns_by_normalized_name(df, normalized_names_to_drop):
    keep_columns = [
        column_name
        for column_name in df.columns
        if normalize_column_name(column_name) not in normalized_names_to_drop
    ]
    return df[keep_columns].copy()


def dataframe_rows_to_text(rows):
    if not rows:
        return "No rows found."
    df = pd.DataFrame(rows)
    if df.empty:
        return "No rows found."
    return df.fillna("").to_string(index=False)


def dataframe_rows_to_html_table(rows):
    if not rows:
        return "<p>No rows found.</p>"

    df = pd.DataFrame(rows).fillna("")
    columns = list(df.columns)
    if not columns:
        return "<p>No rows found.</p>"

    header_html = "".join(
        f"<th style='border: 1px solid #cbd5e1; padding: 8px; background: #e2e8f0; text-align: left;'>{html.escape(str(col))}</th>"
        for col in columns
    )
    body_rows = []
    for _, row in df.iterrows():
        row_html = "".join(
            f"<td style='border: 1px solid #cbd5e1; padding: 6px; vertical-align: top;'>{html.escape(str(row[col]))}</td>"
            for col in columns
        )
        body_rows.append(f"<tr>{row_html}</tr>")
    body_html = "".join(body_rows)

    return f"""
<table style="border-collapse: collapse; width: 100%; font-family: Calibri, Arial, sans-serif; font-size: 12px;">
  <thead>
    <tr>{header_html}</tr>
  </thead>
  <tbody>
    {body_html}
  </tbody>
</table>
"""


def send_via_outlook_mac(to_email_list, cc_email_list, subject, body, html_body, attachment_paths):
    script = """
on run argv
	set toList to item 1 of argv
	set ccList to item 2 of argv
	set theSubject to item 3 of argv
	set theBody to item 4 of argv
	set theHtmlBody to item 5 of argv
	set attachmentPaths to item 6 of argv
	
	try
		tell application "Microsoft Outlook"
			if theHtmlBody is not "" then
				set newMessage to make new outgoing message with properties {subject:theSubject, content:theHtmlBody}
			else
				set newMessage to make new outgoing message with properties {subject:theSubject, content:theBody}
			end if
			tell newMessage
				if toList is not "" then
					set AppleScript's text item delimiters to ","
					repeat with recipientAddress in text items of toList
						set trimmedTo to my trim_edges(recipientAddress as text)
						if trimmedTo is not "" then
							make new to recipient at end of to recipients with properties {email address:{address:trimmedTo}}
						end if
					end repeat
				end if
				
				if ccList is not "" then
					set AppleScript's text item delimiters to ","
					repeat with recipientAddress in text items of ccList
						set trimmedCc to my trim_edges(recipientAddress as text)
						if trimmedCc is not "" then
							make new cc recipient at end of cc recipients with properties {email address:{address:trimmedCc}}
						end if
					end repeat
				end if
				
				if attachmentPaths is not "" then
					set AppleScript's text item delimiters to linefeed
					repeat with attachmentPath in text items of attachmentPaths
						set trimmedAttachment to my trim_edges(attachmentPath as text)
						if trimmedAttachment is not "" then
							try
								set attachmentAlias to (POSIX file trimmedAttachment) as alias
								make new attachment with properties {file:attachmentAlias}
							on error
								make new attachment with properties {file:(POSIX file trimmedAttachment)}
							end try
						end if
					end repeat
				end if
				send
			end tell
		end tell
	on error errMsg number errNum
		error "Outlook AppleScript failed (" & errNum & "): " & errMsg
	end try
end run

on trim_edges(theText)
	set t to theText as text
	repeat while t begins with " " or t begins with tab or t begins with return or t begins with linefeed
		if (length of t) ≤ 1 then return ""
		set t to text 2 thru -1 of t
	end repeat
	repeat while t ends with " " or t ends with tab or t ends with return or t ends with linefeed
		if (length of t) ≤ 1 then return ""
		set t to text 1 thru -2 of t
	end repeat
	return t
end trim_edges
"""
    command = [
        "osascript",
        "-e",
        script,
        ",".join(to_email_list),
        ",".join(cc_email_list),
        subject,
        body,
        html_body,
        "\n".join(attachment_paths),
    ]
    try:
        subprocess.run(command, check=True, capture_output=True, text=True)
    except subprocess.CalledProcessError as exc:
        stderr = (exc.stderr or "").strip()
        stdout = (exc.stdout or "").strip()
        details = stderr or stdout or str(exc)
        raise RuntimeError(details) from exc


def get_smtp_server():
    smtp_host = os.getenv("SMTP_HOST")
    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    smtp_user = os.getenv("SMTP_USER")
    smtp_password = os.getenv("SMTP_PASSWORD")
    smtp_from = os.getenv("SMTP_FROM", smtp_user or "")
    use_starttls = os.getenv("SMTP_STARTTLS", "true").lower() in {"1", "true", "yes"}

    if not smtp_host or not smtp_user or not smtp_password:
        raise RuntimeError(
            "SMTP config missing. Set SMTP_HOST, SMTP_USER, SMTP_PASSWORD "
            "(optional: SMTP_PORT, SMTP_FROM, SMTP_STARTTLS)."
        )

    smtp_server = smtplib.SMTP(smtp_host, smtp_port)
    smtp_server.ehlo()
    if use_starttls:
        smtp_server.starttls()
        smtp_server.ehlo()
    smtp_server.login(smtp_user, smtp_password)
    return smtp_server, smtp_from


def to_excel_bytes(rows, workflow_id):
    output = BytesIO()
    df = pd.DataFrame(rows)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)
        worksheet = writer.book.active

        if workflow_id in RSTC_HIGHLIGHT_WORKFLOWS and len(df.columns) > 0:
            thin_side = Side(style="thin", color="000000")
            thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
            centered = Alignment(horizontal="center", vertical="center")

            # Apply table-like formatting (thin borders + centered text) to header and data cells.
            max_row = len(df.index) + 1
            max_col = len(df.columns)
            for row_idx in range(1, max_row + 1):
                for col_idx in range(1, max_col + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    cell.alignment = centered

            # Auto-fit each column width based on max content length (header + cells).
            for col_idx, column_name in enumerate(df.columns, start=1):
                max_len = len(str(column_name))
                for value in df[column_name].tolist():
                    value_len = len("" if pd.isna(value) else str(value))
                    if value_len > max_len:
                        max_len = value_len
                worksheet.column_dimensions[worksheet.cell(row=1, column=col_idx).column_letter].width = min(
                    max(12, max_len + 2), 80
                )

        if not df.empty and len(df.columns) > 0:
            green_header_fill = PatternFill(
                start_color="A9D08E",
                end_color="A9D08E",
                fill_type="solid",
            )
            for excel_col_idx in range(1, len(df.columns) + 1):
                header_cell = worksheet.cell(row=1, column=excel_col_idx)
                header_cell.fill = green_header_fill
                header_cell.font = Font(bold=True)

        if workflow_id in RSTC_HIGHLIGHT_WORKFLOWS and not df.empty:
            normalized_cols = [normalize_column_name(col) for col in df.columns]
            status_col_indexes = [
                idx
                for idx, normalized_name in enumerate(normalized_cols)
                if normalized_name in {"status", "statuscode"}
            ]
            if status_col_indexes:
                yellow_fill = PatternFill(
                    start_color="FFF59D",
                    end_color="FFF59D",
                    fill_type="solid",
                )
                for df_row_idx, row_values in enumerate(df.itertuples(index=False), start=2):
                    should_highlight = False
                    for col_idx in status_col_indexes:
                        value = row_values[col_idx]
                        text_value = "" if pd.isna(value) else str(value).strip().upper()
                        if text_value not in {"BV", "EV"}:
                            should_highlight = True
                            break
                    if should_highlight:
                        for excel_col_idx in range(1, len(df.columns) + 1):
                            worksheet.cell(row=df_row_idx, column=excel_col_idx).fill = yellow_fill
    return output.getvalue()


def sanitize_for_json(value):
    if pd.isna(value):
        return ""
    if isinstance(value, pd.Timestamp):
        return value.isoformat()
    return value


def sanitize_rows_for_json(rows):
    return [{key: sanitize_for_json(value) for key, value in row.items()} for row in rows]


def create_named_attachment_tempfile(attachment_name, attachment_bytes):
    temp_dir = tempfile.mkdtemp(prefix="excel-emailer-")
    temp_path = os.path.join(temp_dir, attachment_name)
    with open(temp_path, "wb") as temp_file:
        temp_file.write(attachment_bytes)
    return temp_path, temp_dir


def send_email(record, workflow_id, use_outlook_windows, use_outlook_mac, outlook, smtp_server, smtp_from):
    office_name_upper = str(record.get("office_name", "")).strip().upper()
    if office_name_upper == "FREDPEDO":
        appointment_for_subject = format_date_for_display(record.get("appointment_date", ""))
        if appointment_for_subject:
            appointment_for_subject = appointment_for_subject.replace("/", ".")
        else:
            appointment_for_subject = datetime.now().strftime("%m.%d.%Y")
        subject = f"FREDPEDO DENTAL ELIGIBILITY REPORT FOR DOS {appointment_for_subject}"
    elif office_name_upper == "MUSGROVE":
        appointment_for_subject = format_date_for_display(record.get("appointment_date", ""))
        if appointment_for_subject:
            appointment_for_subject = appointment_for_subject.replace("/", "")
        else:
            appointment_for_subject = datetime.now().strftime("%m%d%Y")
        subject = (
            f"MUSGROVE DENTAL ELIGIBLITY AND HISTORIES REPORT FOR DOS {appointment_for_subject}"
        )
    elif workflow_uses_no_info_body_mode(workflow_id):
        subject = f"{record['office_name']} No Information - ({datetime.now().strftime('%m/%d/%Y')})"
    elif workflow_id in EV_RSTC_WORKFLOWS:
        appointment_date = record.get("appointment_date") or datetime.now().strftime("%m/%d/%Y")
        subject = f"{record['office_name']} Dental Eligibility Report - Appt Date ({appointment_date})"
    elif workflow_id in BV_PREV_DAY_WORKFLOWS:
        previous_day = (datetime.now() - timedelta(days=1)).strftime("%m/%d/%Y")
        subject = f"{record['office_name']} BV Report - ({previous_day})"
    else:
        subject = f"BV Report - {record['office_name']}"
    use_inline_no_info = workflow_uses_no_info_body_mode(workflow_id)
    if use_inline_no_info:
        table_text = dataframe_rows_to_text(record["slice_rows"])
        table_html = dataframe_rows_to_html_table(record["slice_rows"])
        body = f"""Hello,

Kindly note that we were unable to locate the insurance information for the patient mentioned below in our system. We request you to kindly share the patient's insurance details via email so that we can verify the coverage and benefits accordingly.

{table_text}

Regards,
Mushtaq Memon
"""
        body_html = f"""
<html>
  <body style="font-family: Calibri, Arial, sans-serif; font-size: 13px; color: #1f2937;">
    <p>Hello,</p>
    <p>Kindly note that we were unable to locate the insurance information for the patient mentioned below in our system. We request you to kindly share the patient's insurance details via email so that we can verify the coverage and benefits accordingly.</p>
    {table_html}
    <p style="margin-top: 16px;">Regards,<br/>Mushtaq Memon</p>
  </body>
</html>
"""
        attachment_bytes = None
        attachment_name = None
    else:
        body = f"""Hi,

Please find attached BV Report for {record['office_name']}.

Regards,
Mushtaq Memon
"""
        body_html = None
        attachment_bytes = to_excel_bytes(record["slice_rows"], workflow_id)
        office_name_for_file = safe_file_name(record["office_name"]) or record.get("safe_name") or "Report"
        office_name_upper = str(record.get("office_name", "")).strip().upper()
        if office_name_upper == "FREDPEDO":
            appointment_for_name = format_date_for_display(record.get("appointment_date", ""))
            if appointment_for_name:
                appointment_for_name = appointment_for_name.replace("/", ".")
            else:
                appointment_for_name = datetime.now().strftime("%m.%d.%Y")
            attachment_name = (
                f"FREDPEDO DENTAL ELIGIBILITY REPORT FOR DOS {appointment_for_name}.xlsx"
            )
        elif office_name_upper == "MUSGROVE":
            appointment_for_name = format_date_for_display(record.get("appointment_date", ""))
            if appointment_for_name:
                appointment_for_name = appointment_for_name.replace("/", "")
            else:
                appointment_for_name = datetime.now().strftime("%m%d%Y")
            attachment_name = (
                f"MUSGROVE DENTAL ELIGIBLITY AND HISTORIES REPORT FOR DOS {appointment_for_name}.xlsx"
            )
        elif workflow_id in BV_PREV_DAY_WORKFLOWS:
            us_date_for_file = (datetime.now() - timedelta(days=1)).strftime("%m-%d-%Y")
            attachment_name = f"{office_name_for_file} {us_date_for_file}.xlsx"
        else:
            us_date_for_file = datetime.now().strftime("%m-%d-%Y")
            attachment_name = f"{office_name_for_file} {us_date_for_file}.xlsx"

    attachments_for_send = []
    if not use_inline_no_info:
        attachments_for_send.append(
            {
                "name": attachment_name,
                "bytes": attachment_bytes,
                "maintype": "application",
                "subtype": "vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            }
        )

    extra_pdf_bytes = record.get("extra_pdf_bytes")
    extra_pdf_name = record.get("extra_pdf_name")
    if (
        should_attach_extra_pdf(workflow_id, record.get("office_name", ""))
        and extra_pdf_bytes
        and extra_pdf_name
    ):
        attachments_for_send.append(
            {
                "name": extra_pdf_name,
                "bytes": extra_pdf_bytes,
                "maintype": "application",
                "subtype": "pdf",
            }
        )

    if use_outlook_windows and outlook is not None:
        temp_paths = []
        temp_dirs = []
        for attachment in attachments_for_send:
            temp_path, temp_dir = create_named_attachment_tempfile(
                attachment["name"], attachment["bytes"]
            )
            temp_paths.append(temp_path)
            temp_dirs.append(temp_dir)
        try:
            mail = outlook.CreateItem(0)
            mail.To = "; ".join(record["to"])
            mail.CC = "; ".join(record["cc"])
            mail.Subject = subject
            if use_inline_no_info and body_html:
                mail.HTMLBody = body_html
            else:
                mail.Body = body
            for temp_path in temp_paths:
                mail.Attachments.Add(temp_path)
            mail.Send()
        finally:
            for temp_dir in temp_dirs:
                try:
                    shutil.rmtree(temp_dir, ignore_errors=True)
                except OSError:
                    pass
    elif use_outlook_mac:
        temp_paths = []
        temp_dirs = []
        for attachment in attachments_for_send:
            temp_path, temp_dir = create_named_attachment_tempfile(
                attachment["name"], attachment["bytes"]
            )
            temp_paths.append(temp_path)
            temp_dirs.append(temp_dir)
        try:
            send_via_outlook_mac(
                record["to"],
                record["cc"],
                subject,
                body,
                body_html or "",
                temp_paths,
            )
        finally:
            for temp_dir in temp_dirs:
                try:
                    shutil.rmtree(temp_dir, ignore_errors=True)
                except OSError:
                    pass
    else:
        message = EmailMessage()
        message["From"] = smtp_from
        message["To"] = ", ".join(record["to"])
        if record["cc"]:
            message["Cc"] = ", ".join(record["cc"])
        message["Subject"] = subject
        message.set_content(body)
        if use_inline_no_info and body_html:
            message.add_alternative(body_html, subtype="html")
        for attachment in attachments_for_send:
            message.add_attachment(
                attachment["bytes"],
                maintype=attachment["maintype"],
                subtype=attachment["subtype"],
                filename=attachment["name"],
            )
        smtp_server.send_message(message)


def build_records(workflow_id, email_bytes, data_bytes):
    # Always use first sheet, and auto-detect if files were uploaded in swapped order.
    first_df = pd.read_excel(BytesIO(email_bytes), sheet_name=0)
    second_df = pd.read_excel(BytesIO(data_bytes), sheet_name=0)
    first_df.columns = first_df.columns.str.strip()
    second_df.columns = second_df.columns.str.strip()

    def has_required_columns(df, required_names):
        try:
            for name in required_names:
                find_column(df, name)
            return True
        except ValueError:
            return False

    email_required = ["Doctor Name", "BV Report Send To", "BV Report Send CC"]
    data_required = ["Office Name"]
    first_is_email = has_required_columns(first_df, email_required)
    second_is_data = has_required_columns(second_df, data_required)
    second_is_email = has_required_columns(second_df, email_required)
    first_is_data = has_required_columns(first_df, data_required)

    if first_is_email and second_is_data:
        email_df, data_df = first_df, second_df
    elif second_is_email and first_is_data:
        email_df, data_df = second_df, first_df
    else:
        raise ValueError(
            "Could not identify Email and Data files. "
            "Email file must contain Doctor Name, BV Report Send To, BV Report Send CC; "
            "Data file must contain Office Name."
        )

    doctor_col = find_column(email_df, "Doctor Name")
    to_col = find_column(email_df, "BV Report Send To")
    cc_col = find_column(email_df, "BV Report Send CC")
    office_col = find_column(data_df, "Office Name")
    no_info_columns = []
    ev_appointment_col = None
    if workflow_uses_no_info_body_mode(workflow_id):
        no_info_columns = [
            find_column(data_df, "Insurance"),
            find_column(data_df, "Policy id"),
            find_column(data_df, "Tel"),
            find_column(data_df, "Subscriber Name"),
            find_column(data_df, "Subscriber DOB"),
        ]
    elif workflow_id in EV_RSTC_WORKFLOWS:
        try:
            ev_appointment_col = find_column(data_df, "Appointment")
        except ValueError:
            try:
                ev_appointment_col = find_column(data_df, "Appoinment Date")
            except ValueError:
                ev_appointment_col = None
    office_names = data_df[office_col].fillna("").astype(str).str.strip()
    doctor_names = email_df[doctor_col].fillna("").astype(str).str.strip()
    doctor_name_set = {name for name in doctor_names if name and name.lower() != "nan"}

    records = []
    skipped_no_doctor = 0
    skipped_no_to = 0
    skipped_no_data = 0

    for _, row in email_df.iterrows():
        doctor = str(row[doctor_col]).strip()
        to_email_list = clean_recipients(row[to_col])
        cc_email_list = clean_recipients(row[cc_col])
        record = {
            "id": str(uuid.uuid4()),
            "office_name": doctor,
            "appointment_date": "",
            "safe_name": "",
            "to": to_email_list,
            "cc": cc_email_list,
            "status": "pending",
            "message": "",
            "row_count": 0,
            "slice_rows": [],
            "can_send": False,
            "extra_pdf_name": "",
            "extra_pdf_bytes": b"",
        }

        if not doctor or doctor.lower() == "nan":
            skipped_no_doctor += 1
            continue

        if not to_email_list:
            skipped_no_to += 1
            continue

        filtered_df = data_df[office_names == doctor]
        if workflow_uses_no_info_body_mode(workflow_id):
            filtered_df = filter_no_info_rows(filtered_df, no_info_columns)
        if filtered_df.empty:
            skipped_no_data += 1
            continue

        formatted_filtered_df = format_date_columns_in_df(filtered_df)
        if doctor.strip().upper() == "FREDPEDO":
            formatted_filtered_df = transform_fredpedo_slice_df(formatted_filtered_df)
            formatted_filtered_df = format_date_columns_in_df(formatted_filtered_df)
        elif workflow_id in EV_RSTC_WORKFLOWS:
            formatted_filtered_df = drop_columns_by_normalized_name(
                formatted_filtered_df,
                EV_EXCLUDED_OUTPUT_COLUMNS,
            )
        clean_name = safe_file_name(doctor.replace(".", "").replace("&", "and")) or "Report"

        if workflow_id in EV_RSTC_WORKFLOWS and ev_appointment_col and ev_appointment_col in formatted_filtered_df.columns:
            grouped = formatted_filtered_df.groupby(ev_appointment_col, dropna=False, sort=False)
            for appointment_value, appointment_df in grouped:
                appointment_date = format_date_for_display(appointment_value)
                grouped_record = record.copy()
                grouped_record["id"] = str(uuid.uuid4())
                grouped_record["can_send"] = True
                grouped_record["appointment_date"] = appointment_date
                grouped_record["safe_name"] = (
                    safe_file_name(f"{clean_name} {appointment_date}") or clean_name
                )
                grouped_record["row_count"] = len(appointment_df)
                grouped_record["slice_rows"] = appointment_df.fillna("").to_dict(orient="records")
                records.append(grouped_record)
        else:
            record["can_send"] = True
            record["safe_name"] = clean_name
            record["row_count"] = len(formatted_filtered_df)
            record["slice_rows"] = formatted_filtered_df.fillna("").to_dict(orient="records")
            records.append(record)

    unmatched_data_df = data_df[(~office_names.isin(doctor_name_set)) & (office_names != "")]
    if workflow_uses_no_info_body_mode(workflow_id):
        unmatched_data_df = filter_no_info_rows(unmatched_data_df, no_info_columns)
    unmatched_data_df = format_date_columns_in_df(unmatched_data_df)
    if workflow_id in EV_RSTC_WORKFLOWS:
        unmatched_data_df = drop_columns_by_normalized_name(
            unmatched_data_df,
            EV_EXCLUDED_OUTPUT_COLUMNS,
        )
    unmatched_rows = unmatched_data_df.fillna("").to_dict(orient="records")
    if unmatched_rows:
        records.append(
            {
                "id": str(uuid.uuid4()),
                "office_name": "Miscellaneous (Non Matching)",
                "safe_name": "Misc-Non-Matching-Report",
                "to": MISC_TO_RECIPIENTS.copy(),
                "cc": MISC_CC_RECIPIENTS.copy(),
                "status": "pending",
                "message": "All non-matching records grouped here.",
                "row_count": len(unmatched_rows),
                "slice_rows": unmatched_rows,
                "can_send": True,
            }
        )

    summary = {
        "rows": len(email_df),
        "sent": 0,
        "failed": 0,
        "skipped_no_doctor": skipped_no_doctor,
        "skipped_no_to": skipped_no_to,
        "skipped_no_data": skipped_no_data,
    }
    return records, summary


def send_records(workflow_id, record_ids=None):
    state = STATE_BY_WORKFLOW[workflow_id]
    use_outlook_windows = os.name == "nt" and win32 is not None
    use_outlook_mac = sys.platform == "darwin"
    com_initialized = False

    if use_outlook_windows and pythoncom is not None:
        pythoncom.CoInitialize()
        com_initialized = True

    outlook = win32.Dispatch("outlook.application") if use_outlook_windows else None
    smtp_server = None
    smtp_from = ""

    if not use_outlook_windows and not use_outlook_mac:
        smtp_server, smtp_from = get_smtp_server()

    try:
        sent_count = 0
        failed_count = 0
        for record in state["records"]:
            if record_ids and record["id"] not in record_ids:
                continue

            if not record["can_send"]:
                continue

            try:
                send_email(
                    record,
                    workflow_id,
                    use_outlook_windows,
                    use_outlook_mac,
                    outlook,
                    smtp_server,
                    smtp_from,
                )
                record["status"] = "sent"
                record["message"] = "Sent successfully"
                sent_count += 1
                print(f"[SENT] {record['office_name']} -> {', '.join(record['to'])}")
            except Exception as e:
                record["status"] = "failed"
                record["message"] = str(e)
                failed_count += 1
                print(f"[FAIL] {record['office_name']}: {e}")

        state["summary"]["sent"] += sent_count
        state["summary"]["failed"] += failed_count
        return state["summary"]
    finally:
        if smtp_server is not None:
            smtp_server.quit()
        if com_initialized and pythoncom is not None:
            pythoncom.CoUninitialize()


@app.get("/")
def home():
    workflow_id = get_workflow_or_default(request.args.get("workflow", DEFAULT_WORKFLOW))
    state = STATE_BY_WORKFLOW[workflow_id]
    return render_template(
        "index.html",
        records=state["records"],
        summary=state["summary"],
        uploaded=state["uploaded"],
        workflows=WORKFLOWS,
        current_workflow=workflow_id,
        current_workflow_label=WORKFLOW_MAP[workflow_id],
    )


@app.post("/upload/<workflow_id>")
def upload_files(workflow_id):
    workflow_id = get_workflow_or_default(workflow_id)
    state = STATE_BY_WORKFLOW[workflow_id]
    try:
        email_upload = request.files.get("email_file")
        data_upload = request.files.get("data_file")
        if not email_upload or not data_upload:
            flash("Please upload both files.", "error")
            return redirect(url_for("home"))

        email_bytes = email_upload.read()
        data_bytes = data_upload.read()
        records, summary = build_records(workflow_id, email_bytes, data_bytes)
        state["records"] = records
        state["summary"] = summary
        state["uploaded"] = True

        if not records:
            flash(
                "Upload completed, but no sendable records were found after matching/filtering.",
                "error",
            )
        else:
            flash("Files uploaded and sliced reports generated.", "success")
        return redirect(url_for("home", workflow=workflow_id))
    except Exception as e:
        flash(f"Upload failed: {e}", "error")
        return redirect(url_for("home", workflow=workflow_id))


@app.post("/send-all/<workflow_id>")
def send_all(workflow_id):
    workflow_id = get_workflow_or_default(workflow_id)
    state = STATE_BY_WORKFLOW[workflow_id]
    if not state["uploaded"]:
        flash("Upload both files first.", "error")
        return redirect(url_for("home", workflow=workflow_id))
    try:
        send_records(workflow_id)
        flash("Send All completed.", "success")
    except Exception as e:
        flash(f"Send All failed: {e}", "error")
    return redirect(url_for("home", workflow=workflow_id))


@app.post("/upload-extra-pdf/<workflow_id>/<record_id>")
def upload_extra_pdf(workflow_id, record_id):
    workflow_id = get_workflow_or_default(workflow_id)
    state = STATE_BY_WORKFLOW[workflow_id]
    record = next((r for r in state["records"] if r["id"] == record_id), None)
    if not record:
        flash("Record not found.", "error")
        return redirect(url_for("home", workflow=workflow_id))

    if not should_attach_extra_pdf(workflow_id, record.get("office_name", "")):
        flash("Extra PDF upload is not enabled for this record.", "error")
        return redirect(url_for("home", workflow=workflow_id))

    pdf_upload = request.files.get("extra_pdf_file")
    if not pdf_upload or not pdf_upload.filename:
        flash("Please choose a PDF file to upload.", "error")
        return redirect(url_for("home", workflow=workflow_id))

    filename = pdf_upload.filename.strip()
    if not filename.lower().endswith(".pdf"):
        flash("Only PDF files are allowed.", "error")
        return redirect(url_for("home", workflow=workflow_id))

    file_bytes = pdf_upload.read()
    if not file_bytes:
        flash("Uploaded PDF is empty.", "error")
        return redirect(url_for("home", workflow=workflow_id))

    safe_name = safe_file_name(filename.rsplit(".", 1)[0]) or "attachment"
    record["extra_pdf_name"] = f"{safe_name}.pdf"
    record["extra_pdf_bytes"] = file_bytes
    flash(f"PDF attached for {record['office_name']}.", "success")
    return redirect(url_for("home", workflow=workflow_id))


@app.post("/reset/<workflow_id>")
def reset_app(workflow_id):
    workflow_id = get_workflow_or_default(workflow_id)
    try:
        reset_state(workflow_id)
        flash("App reset completed.", "success")
    except Exception as e:
        flash(f"Reset failed: {e}", "error")
    return redirect(url_for("home", workflow=workflow_id))


@app.post("/send/<workflow_id>/<record_id>")
def send_one(workflow_id, record_id):
    workflow_id = get_workflow_or_default(workflow_id)
    state = STATE_BY_WORKFLOW[workflow_id]
    if not state["uploaded"]:
        flash("Upload both files first.", "error")
        return redirect(url_for("home", workflow=workflow_id))

    record = next((r for r in state["records"] if r["id"] == record_id), None)
    if not record:
        flash("Record not found.", "error")
        return redirect(url_for("home", workflow=workflow_id))
    if not record["can_send"]:
        flash("This record cannot be sent.", "error")
        return redirect(url_for("home", workflow=workflow_id))

    try:
        send_records(workflow_id, record_ids={record_id})
        flash(f"Sent: {record['office_name']}", "success")
    except Exception as e:
        flash(f"Failed to send {record['office_name']}: {e}", "error")
    return redirect(url_for("home", workflow=workflow_id))


@app.get("/preview/<workflow_id>/<record_id>")
def preview_slice(workflow_id, record_id):
    workflow_id = get_workflow_or_default(workflow_id)
    state = STATE_BY_WORKFLOW[workflow_id]
    record = next((r for r in state["records"] if r["id"] == record_id), None)
    if not record or not record.get("slice_rows"):
        return jsonify({"ok": False, "error": "Sliced file not found."}), 404

    try:
        df = pd.DataFrame(record["slice_rows"])
        rows = sanitize_rows_for_json(df.head(200).to_dict(orient="records"))
        highlight_flags = []
        if workflow_id in RSTC_HIGHLIGHT_WORKFLOWS and not df.empty:
            normalized_cols = [normalize_column_name(col) for col in df.columns]
            status_col_indexes = [
                idx
                for idx, normalized_name in enumerate(normalized_cols)
                if normalized_name in {"status", "statuscode"}
            ]
            if status_col_indexes:
                preview_df = df.head(200)
                for row_values in preview_df.itertuples(index=False):
                    should_highlight = False
                    for col_idx in status_col_indexes:
                        value = row_values[col_idx]
                        text_value = "" if pd.isna(value) else str(value).strip().upper()
                        if text_value not in {"BV", "EV"}:
                            should_highlight = True
                            break
                    highlight_flags.append(should_highlight)
            else:
                highlight_flags = [False] * len(rows)
        else:
            highlight_flags = [False] * len(rows)
        return jsonify(
            {
                "ok": True,
                "office_name": record["office_name"],
                "columns": list(df.columns),
                "rows": rows,
                "highlight_flags": highlight_flags,
            }
        )
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5004, debug=True)